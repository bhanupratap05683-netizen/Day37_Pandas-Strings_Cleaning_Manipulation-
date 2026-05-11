"""
Day 37 - String Cleaning & Manipulation
Topic: strip, replace, split, case conversion, regex basics
Dataset: Messy financial transaction records + employee data
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: PYTHON STRING METHODS (single string — no pandas yet)
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("SECTION 1: Core Python String Methods")
print("=" * 60)

raw = "  Reliance Industries  "

# strip() — removes whitespace from BOTH sides
print(raw.strip())          # "Reliance Industries"

# lstrip() / rstrip() — removes from left OR right only
print(raw.lstrip())         # "Reliance Industries  "
print(raw.rstrip())         # "  Reliance Industries"

# Case conversion
word = "hdfc bank"
print(word.upper())         # "HDFC BANK"
print(word.lower())         # "hdfc bank"
print(word.title())         # "Hdfc Bank"
print(word.capitalize())    # "Hdfc bank"  ← only first letter

# replace() — swap one substring for another
amount = "Rs. 45,000.50"
amount = amount.replace("Rs. ", "").replace(",", "")
print(float(amount))        # 45000.5

# split() — break string into a list
txn_id = "TXN-001"
parts = txn_id.split("-")
print(parts)                # ['TXN', '001']
print(parts[0])             # 'TXN'
print(parts[1])             # '001'

# startswith() / endswith()
email = "rahul.sharma@firm.com"
print(email.startswith("rahul"))     # True
print(email.endswith(".com"))        # True

# in — check if substring exists
note = "URGENT: verify before eod"
print("URGENT" in note)              # True

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: REGEX BASICS (re module)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("SECTION 2: Regex Basics with re Module")
print("=" * 60)

# What is regex? A pattern language to search/replace inside strings.
# Common patterns:
#   \d    = any digit (0-9)
#   \D    = any non-digit
#   \s    = any whitespace (space, tab, newline)
#   \S    = any non-whitespace
#   \w    = any word character (letters, digits, _)
#   +     = one or more of the previous
#   *     = zero or more
#   []    = character set  e.g. [^0-9] means NOT a digit
#   ^     = NOT (inside []) OR start of string (outside [])

# re.sub(pattern, replacement, string) — like replace() but uses patterns
phone = "+91-9876543210"
digits_only = re.sub(r"[^0-9]", "", phone)   # remove everything that's NOT a digit
print(digits_only)           # 9876543210

# Clean amount: strip prefix first, then remove commas
amount_messy = "Rs. 45,000.50"
amount_clean = re.sub(r"^[A-Za-z.\s]+", "", amount_messy).replace(",", "")
print(float(amount_clean))   # 45000.5

# re.search(pattern, string) — find first match
email_raw = "ankit.gupta @firm.com"
space_found = re.search(r"\s", email_raw)
print(bool(space_found))     # True  ← space exists inside email

# re.findall(pattern, string) — find ALL matches as a list
text = "TXN-001 TXN-004 TXN-007"
ids = re.findall(r"TXN-\d+", text)
print(ids)                   # ['TXN-001', 'TXN-004', 'TXN-007']

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: PANDAS .str ACCESSOR (vectorised — applies to whole column)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("SECTION 3: pandas .str Accessor")
print("=" * 60)

# Load messy data
df_txn = pd.read_excel("Day37_String_Cleaning_Input.xlsx", sheet_name="Messy_Transactions", dtype=str)
df_emp = pd.read_excel("Day37_String_Cleaning_Input.xlsx", sheet_name="Employee_Data", dtype=str)

print("\n--- Raw Transactions (first 3 rows) ---")
print(df_txn.head(3).to_string(index=False))

# ── Clean Transactions ────────────────────────────────────────────────────────
# 1. Transaction_ID: strip whitespace
df_txn["Transaction_ID"] = df_txn["Transaction_ID"].str.strip()

# 2. Company_Name: strip + title case
df_txn["Company_Name"] = df_txn["Company_Name"].str.strip().str.title()

# 3. Transaction_Type: strip + upper + remove extra internal spaces
df_txn["Transaction_Type"] = df_txn["Transaction_Type"].str.strip().str.upper()

# 4. Amount_Raw → clean float
#    "Rs. 45,000.50" → 45000.50
#    Step: remove everything that's not digit or dot using regex
df_txn["Amount_Clean"] = (
    df_txn["Amount_Raw"]
    .str.replace(r"^[A-Za-z.\s]+", "", regex=True)   # remove text prefix (Rs, Rs.)
    .str.replace(",", "", regex=False)                  # remove thousands commas
    .astype(float)
)

# 5. Analyst_Note: strip + title case + flag URGENT
df_txn["Note_Cleaned"] = df_txn["Analyst_Note"].str.strip().str.capitalize()
df_txn["Is_Urgent"]    = df_txn["Analyst_Note"].str.contains("URGENT", case=False, na=False)

print("\n--- Cleaned Transactions ---")
print(df_txn[["Transaction_ID", "Company_Name", "Transaction_Type",
              "Amount_Clean", "Note_Cleaned", "Is_Urgent"]].to_string(index=False))

# ── Clean Employee Data ───────────────────────────────────────────────────────
# 1. Emp_ID: strip
df_emp["Emp_ID"] = df_emp["Emp_ID"].str.strip()

# 2. Full_Name: strip + title case
df_emp["Full_Name"] = df_emp["Full_Name"].str.strip().str.title()

# 3. Department: strip + title case
df_emp["Department"] = df_emp["Department"].str.strip().str.title()

# 4. Email: strip + lower + remove internal spaces
df_emp["Email_Clean"] = (
    df_emp["Email_Raw"]
    .str.strip()
    .str.lower()
    .str.replace(r"\s+", "", regex=True)   # remove all whitespace inside email
)

# 5. Phone: extract digits only → format as 10-digit (drop country code if present)
df_emp["Phone_Clean"] = df_emp["Phone_Raw"].str.replace(r"[^0-9]", "", regex=True)
# If 12 digits, it has 91 prefix → take last 10
df_emp["Phone_Clean"] = df_emp["Phone_Clean"].apply(
    lambda x: x[-10:] if len(x) >= 10 else x
)

# 6. City: strip + title case
df_emp["City"] = df_emp["City"].str.strip().str.title()

print("\n--- Cleaned Employees ---")
print(df_emp[["Emp_ID", "Full_Name", "Department", "Email_Clean", "Phone_Clean", "City"]].to_string(index=False))

# ── Summary Statistics ────────────────────────────────────────────────────────
print("\n--- Summary ---")
print(f"Total Transactions : {len(df_txn)}")
print(f"Purchase Count     : {(df_txn['Transaction_Type'] == 'PURCHASE').sum()}")
print(f"Sale Count         : {(df_txn['Transaction_Type'] == 'SALE').sum()}")
print(f"Total Amount (Rs.) : {df_txn['Amount_Clean'].sum():,.2f}")
print(f"Avg Amount (Rs.)   : {df_txn['Amount_Clean'].mean():,.2f}")
print(f"Urgent Flags       : {df_txn['Is_Urgent'].sum()}")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: EXPORT CLEANED DATA TO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("SECTION 4: Exporting Cleaned Data to Excel")
print("=" * 60)

output_path = "Day37_String_Cleaning_Output.xlsx"

# Drop raw/intermediate columns before export
txn_export = df_txn.drop(columns=["Amount_Raw", "Analyst_Note"])
emp_export  = df_emp.drop(columns=["Email_Raw", "Phone_Raw"])

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    txn_export.to_excel(writer, sheet_name="Clean_Transactions", index=False)
    emp_export.to_excel(writer, sheet_name="Clean_Employees",    index=False)

# Apply formatting
wb = load_workbook(output_path)

header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
green_fill   = PatternFill("solid", start_color="1E6B3C")   # green = cleaned data
blue_fill    = PatternFill("solid", start_color="1F4E79")
alt_fill     = PatternFill("solid", start_color="E8F5E9")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    fill = green_fill if "Transaction" in ws.title else blue_fill
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 28)
    ws.row_dimensions[1].height = 22

wb.save(output_path)
print(f"Cleaned data saved → {output_path}")
print("\nDay 37 Complete!")
